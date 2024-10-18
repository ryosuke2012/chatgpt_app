import os
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

import openpyxl

base_dir = Path(__file__).parent
excel_path = base_dir / 'chat_log.xlsx'

def is_output_open_excel() -> bool:
    """
    Excelファイルが開かれているかどうかを判定する
    :return: 開かれているかの真偽値
    """

    # Windows
    if os.name == "nt":
        try:
            with excel_path,open("r+b"):
                return False
        except PermissionError:
            return True
        except FileNotFoundError:
            return False

    # Mac
    if os.name == "posix":
        if excel_path.exists():
            result = subprocess.run(["lsof", excel_path], stdout=subprocess.PIPE)
            return bool(result.stdout)
        return False

def load_or_create_workbook() -> tuple[openpyxl.workbook, bool]:
    """
    Excelファイルを読み込むか、存在しない場合は作成する
    :returns: ワークブックオブジェクトと、ファイルが作成されたかどうかのフラグ
    """

    # ファイルの存在確認
    if excel_path.exists():
        # ファイルを読み込んで返す
        wb = openpyxl.load_workbook(excel_path)
        return wb, False
    else:
        # ファイルを作成して返す
        wb = openpyxl.Workbook()
        return wb, True

def create_worksheet(title: str, wb: openpyxl.Workbook, is_new: bool):
    """
    シートを作成してシート名に使えない文字を除去したうえでシート名を変更して返す
    :param title: シート名(プロンプトの要約)
    :param wb: 対象になるワークブックオブジェクト
    :param is_new: ワークブックが新規作成されたかどうかのフラグ
    :return: ワークシートオブジェクト
    """

    # シート名に使えない文字を除去
    trimmed_title = trim_invalid_chars(title)

    if is_new:
        # アクティブシート(Sheet)を取得
        ws = wb.active
        ws.title = trimmed_title
    else:
        # シートを追加
        ws = wb.create_sheet(title=trimmed_title)
        wb.move_sheet(ws, offset=-len(wb.worksheets)+1)
        wb.active = ws
    return ws

def trim_invalid_chars(title: str) -> str:
    """
    シート名に使えない文字を除去する
    :param title: シート名
    :return: 除去後のシート名
    """

    invalid_chars = ["/", "\\", "?", "*", "[", "]"]
    for char in invalid_chars:
        title = title.replace(char,"")
    return title

def header_formatting(ws):
    """
    出力対象のワークシートにヘッダーを設定する
    :param ws: 出力対処のワークシート
    """

    # A1セルに出力時の日時を書き込みフォントを設定
    datetime_cell = ws["A1"]
    datetime_cell.value = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    datetime_cell.font = Font(name="メイリオ")

    # A2セルに「ロール」B2セルに「発言内容」と書き込み、フォントとセルの色を設定
    role_header_cell = ws["A2"]
    content_header_cell = ws["B2"]

    # セルに値を設定
    role_header_cell.value = "ロール"
    content_header_cell.value = "発言内容"

    # フォントを設定
    header_font_style = Font(name="メイリオ", bold=True, color="FFFFFF")
    role_header_cell.font = header_font_style
    content_header_cell.font = header_font_style

    # セルの色を設定
    header_color = PatternFill(fill_type="solid", fgColor="156B31")
    role_header_cell.fill = header_color
    content_header_cell.fill = header_color

    # セルの幅を調整
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 168

is_output = is_output_open_excel()
# print(is_output)

workbook, is_created = load_or_create_workbook()
worksheet = create_worksheet("test/\\?[]", workbook, is_created)
header_formatting(worksheet)
workbook.save(excel_path)
print(worksheet.title)
